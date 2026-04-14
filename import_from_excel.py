# import_from_excel.py
#
# Run this ONCE to import your existing Excel tracker into the SQLite database.
# After running it, all 20 applications will appear in your Streamlit app.
#
# Usage:
#   python import_from_excel.py
#
# Make sure this file is in the same folder as app.py and database.py.
# Also make sure your Excel file is in the same folder, or update EXCEL_PATH below.

import sqlite3
from openpyxl import load_workbook
from datetime import datetime

# ── CONFIG ────────────────────────────────────────────────────────────────────
# Update this if your Excel file is in a different location
EXCEL_PATH = "Akhil_JobApplicationTracker.xlsx"
DB_PATH    = "jobs.db"
# ─────────────────────────────────────────────────────────────────────────────

# Map the Excel "Site" column values to the platform names used in the app
PLATFORM_MAP = {
    "Reed":     "Reed",
    "CWJobs":   "CWJobs",
    "Indeed":   "Indeed",
    "Glassdoor":"Glassdoor",
    "NHS Jobs": "NHS Jobs",
    "LinkedIn": "LinkedIn",
}

# Map Excel status values to app status values (just in case they differ)
STATUS_MAP = {
    "Applied":        "Applied",
    "Recruiter Call": "Recruiter Call",
    "Interview":      "Interview",
    "Assessment":     "Assessment",
    "Offer":          "Offer",
    "Rejected":       "Rejected",
    "Withdrawn":      "Withdrawn",
}

def convert_date(raw):
    """Convert date from Excel format (DD/MM/YYYY) to app format (YYYY-MM-DD)."""
    if raw is None:
        return str(datetime.today().date())
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    # Handle string dates like "10/04/2026"
    try:
        return datetime.strptime(str(raw), "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return str(datetime.today().date())

def init_db(conn):
    """Create the applications table if it doesn't exist."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS applications (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            company      TEXT NOT NULL,
            role         TEXT NOT NULL,
            location     TEXT,
            salary       TEXT,
            platform     TEXT,
            status       TEXT NOT NULL DEFAULT 'Applied',
            date_applied TEXT NOT NULL,
            job_url      TEXT,
            notes        TEXT,
            created_at   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()

def import_data():
    # Load the Excel file
    print(f"📂 Reading: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["Applications"]  # Use the Applications sheet

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]  # First row is the header
    data   = rows[1:] # Everything after is actual data

    print(f"📊 Found {len(data)} applications to import\n")

    # Connect to (or create) the SQLite database
    conn = sqlite3.connect(DB_PATH)
    init_db(conn)

    # Check if there's already data to avoid duplicates
    existing = conn.execute("SELECT COUNT(*) FROM applications").fetchone()[0]
    if existing > 0:
        print(f"⚠️  Database already has {existing} record(s).")
        answer = input("Do you want to import anyway? This may create duplicates. (yes/no): ").strip().lower()
        if answer != "yes":
            print("❌ Import cancelled.")
            conn.close()
            return

    imported = 0
    skipped  = 0

    for row in data:
        # Unpack each column from the Excel row
        # Order: #, Role, Company, Match%, Location, Salary, Site, Applied Date, Status, Notes
        _, role, company, match_pct, location, salary, site, date_applied, status, notes = row

        # Skip completely empty rows
        if not company and not role:
            skipped += 1
            continue

        # Clean up values — replace None with empty string
        company      = str(company or "").strip()
        role         = str(role or "").strip()
        location     = str(location or "").strip()
        salary       = str(salary or "").strip()
        platform     = PLATFORM_MAP.get(str(site or "").strip(), str(site or "").strip())
        status       = STATUS_MAP.get(str(status or "Applied").strip(), "Applied")
        date_applied = convert_date(date_applied)
        notes        = str(notes or "").strip()

        # Add match % to notes if it exists — useful context
        if match_pct:
            notes = f"Match: {match_pct}%" + (f" | {notes}" if notes else "")

        conn.execute("""
            INSERT INTO applications
                (company, role, location, salary, platform, status, date_applied, job_url, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (company, role, location, salary, platform, status, date_applied, "", notes))

        print(f"  ✅ Imported: {role} @ {company}")
        imported += 1

    conn.commit()
    conn.close()

    print(f"\n🎉 Done! {imported} applications imported, {skipped} skipped.")
    print("👉 Now run: python -m streamlit run app.py")

if __name__ == "__main__":
    import_data()
